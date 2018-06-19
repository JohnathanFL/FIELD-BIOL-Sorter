#!/bin/env python3

# Written by: Johnathan Lee (JFL747076@gmail.com)
# 12/02/17 for Field Biol

# Run in a folder with a spreadsheet named Database.xlsx containing the standard
# spreadsheet setup and a folder named "Unsorted". Inside that folder, have a
# folder for each of the cameras named ONLY the number of the camera (i.e 6, 11, etc).

# The spreadsheet should have column G follow the format "Cam X" or "Cam XX",
# where X or XX is the camera number. Column A should be 8 digits long.

# The script will save an updated version of the spreadsheet in a file named
# "FixedDatabase.xlsx", and move the photos to another directory named "Sorted",
# with subdirectories for each species.

import os
from openpyxl import Workbook, load_workbook
import shutil

# For copyfile and mkdir
cwd = os.getcwd()
print("Working in " + cwd)

# Load the spreadsheet
db = load_workbook("Database.xlsx")
ws = db.worksheets[0]

# Get the upper bound of the spreadsheet
try:
  upperbound = int(input("Please enter the number of the last row in the spreadsheet: "))
except:
  print("Error: Please enter ONLY numbers (1, 2, 500, etc).")
  exit(0)
  
# Loop through the rows
for i in range(2, upperbound):
    # Get the current id
    idField= ws['A' + str(i)].value

    # Make sure it's not empty first
    if idField != None:
        curID = str(int(idField)).zfill(8)            # Get the full ID as a str
        camNum = str(ws['G' + str(i)].value)          # Get the camera number
        species = str(ws['H' + str(i)].value).title() # Get the species
        camNum = camNum.replace("Cam ", "")           # "Cam 6 " -> "6"
        
        curFileName = curID + ".JPG"               # Derive the current file's name
        newFileName = camNum.zfill(2) + curFileName# Derive the new file's name
        
        #print(str(int(curID)) + " on " + camNum + " was a(n) " + species)
        
        # If the current species' folder doesn't exist, make it.
        if not os.path.isdir(os.path.join(cwd, "Sorted", species)):
            os.mkdir(os.path.join(cwd, "Sorted", species))
            
        # Try to copy the file over. If we can't, just output an error message.
        try:
            shutil.copyfile(os.path.join(cwd, "Unsorted", camNum, curFileName), os.path.join(cwd, "Sorted", species, newFileName))
            ws['A' + str(i)].value = str(camNum.zfill(2) + curID) # Update the spreadsheet
        except:
            print("Failed to find file " + str(camNum) + "/" + curFileName)

# Save the updated spreadsheet.
db.save("FixedDatabase.xlsx")
